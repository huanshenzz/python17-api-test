#!/usr/bin/python3
# -*- coding: utf-8 -*-
# Name: mu_excel2
# Author: 郑
# Time: 2019/5/31
import openpyxl
from openpyxl import load_workbook
class ExcelV2:
    def __init__(self,file_path,sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def get_init_data(self):
        wb =load_workbook(self.file_path)
        sheet =wb['init']
        mobile = sheet.cell(1,2).value
        return mobile

    def updata_init_data(self,value):
        wb = load_workbook(self.file_path)
        sheet = wb['init']
        sheet.cell(1,2).value = value
        wb.save(self.file_path)


    def read_excel(self):
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook[self.sheet_name]
        cases = []  # 所有的cases数据
        max_row = sheet.max_row  # 获取最大行
        max_column = sheet.max_column  # 获取最大列

        # 双循环
        for r in range(2, max_row + 1):  # 遍历行
            case = {}
            for j in range(1, max_column + 1):
                key = sheet.cell(row=1, column=j).value  # 遍历列
                case[key] = sheet.cell(row=r, column=j).value

            cases.append(case)  # 将一行数据放到列表

        workbook.save(self.file_path)
        workbook.close()
        return cases

    def write_result(self, row, actual, result):
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook[self.sheet_name]
        sheet.cell(row=row, column=7).value = actual
        sheet.cell(row=row, column=8).value = result
        workbook.save(self.file_path)
        workbook.close()


